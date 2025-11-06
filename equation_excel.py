import pandas as pd
import json
import os
from typing import Dict, List, Tuple, Any, Optional
from datetime import datetime
import tempfile

class EquationExcelProcessor:
    """Excel Processor for Equation Mode - ConvertKeylogApp API"""
    
    def __init__(self, config: Dict = None):
        self.config = config or {}
        self.mapping = self._load_equation_mapping()
        self.large_file_threshold_mb = 20
        self.large_file_threshold_rows = 10000
    
    def _load_equation_mapping(self) -> Dict:
        """Load Excel mapping configuration for equation mode"""
        return {
            "Hệ phương trình 2 ẩn": {
                "required_columns": ["a11", "a12", "c1", "a21", "a22", "c2"],
                "variables_count": 2,
                "equations_count": 2
            },
            "Hệ phương trình 3 ẩn": {
                "required_columns": ["a11", "a12", "a13", "c1", "a21", "a22", "a23", "c2", "a31", "a32", "a33", "c3"],
                "variables_count": 3,
                "equations_count": 3
            },
            "Hệ phương trình 4 ẩn": {
                "required_columns": ["a11", "a12", "a13", "a14", "c1", "a21", "a22", "a23", "a24", "c2", 
                                   "a31", "a32", "a33", "a34", "c3", "a41", "a42", "a43", "a44", "c4"],
                "variables_count": 4,
                "equations_count": 4
            }
        }
    
    def is_large_file(self, file_path: str) -> Tuple[bool, Dict[str, Any]]:
        """Check if file is too large for normal processing"""
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            # Quick row count estimation using openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            estimated_rows = ws.max_row - 1 if hasattr(ws, 'max_row') else 0
            wb.close()
            
            is_large = (file_size_mb > self.large_file_threshold_mb or 
                       estimated_rows > self.large_file_threshold_rows)
            
            return is_large, {
                'file_size_mb': file_size_mb,
                'estimated_rows': estimated_rows,
                'recommended_processor': 'large_file' if is_large else 'normal',
                'recommended_chunk_size': self._estimate_optimal_chunksize(file_size_mb)
            }
            
        except Exception as e:
            return False, {'error': f'Không thể phân tích file: {str(e)}'}
    
    def _estimate_optimal_chunksize(self, file_size_mb: float) -> int:
        """Estimate optimal chunk size based on file size"""
        if file_size_mb < 1:
            return 1000
        elif file_size_mb < 10:
            return 500
        elif file_size_mb < 50:
            return 250
        else:
            return 100
    
    def read_excel_data(self, file_path: str) -> pd.DataFrame:
        """Read Excel file and normalize data"""
        try:
            # Check if this is a large file
            is_large, file_info = self.is_large_file(file_path)
            
            if is_large:
                raise Exception(
                    f"File quá lớn cho phương thức thông thường!\n"
                    f"Kích thước: {file_info.get('file_size_mb', 0):.1f}MB\n"
                    f"Dòng ước tính: {file_info.get('estimated_rows', 0):,}\n\n"
                    f"Vui lòng sử dụng chế độ xử lý file lớn."
                )
            
            df = pd.read_excel(file_path)
            # Normalize column names (remove extra spaces)
            df.columns = df.columns.str.strip()
            return df
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel: {str(e)}")
    
    def read_excel_data_chunked(self, file_path: str, chunksize: int = 1000):
        """Read Excel data in chunks for large files"""
        try:
            is_large, file_info = self.is_large_file(file_path)
            
            if is_large:
                # Use openpyxl streaming for very large files
                print(f"🔥 Large file detected - using streaming processor")
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                
                # Get header
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                columns = [str(cell) for cell in header_row if cell is not None]
                
                # Process in chunks
                rows_data = []
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    if i > 0 and i % chunksize == 0:
                        # Yield chunk
                        chunk_df = pd.DataFrame(rows_data, columns=columns)
                        chunk_df.columns = chunk_df.columns.str.strip()
                        yield chunk_df
                        rows_data = []
                    
                    row_dict = [str(cell) if cell is not None else "" for cell in row]
                    rows_data.append(row_dict)
                
                # Yield remaining data
                if rows_data:
                    chunk_df = pd.DataFrame(rows_data, columns=columns)
                    chunk_df.columns = chunk_df.columns.str.strip()
                    yield chunk_df
                
                wb.close()
            else:
                # Use pandas chunking for smaller files
                df = pd.read_excel(file_path)
                df.columns = df.columns.str.strip()
                # Yield chunks
                for i in range(0, len(df), chunksize):
                    yield df.iloc[i:i + chunksize]
                    
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel theo chunk: {str(e)}")
    
    def validate_excel_structure(self, df: pd.DataFrame, operation: str) -> Tuple[bool, List[str]]:
        """Validate Excel structure against selected equation operation"""
        missing_columns = []
        
        if operation not in self.mapping:
            return False, [f"Operation '{operation}' không được hỗ trợ"]
        
        required_cols = self.mapping[operation]['required_columns']
        for col in required_cols:
            if col not in df.columns:
                missing_columns.append(col)
        
        return len(missing_columns) == 0, missing_columns
    
    def extract_equation_data(self, row: pd.Series, operation: str) -> List[str]:
        """Extract equation coefficients from Excel row"""
        operation_config = self.mapping.get(operation, {})
        variables_count = operation_config.get('variables_count', 2)
        equations_count = operation_config.get('equations_count', 2)
        
        equations = []
        
        for eq_idx in range(1, equations_count + 1):
            coefficients = []
            
            # Extract coefficients for this equation
            for var_idx in range(1, variables_count + 1):
                coeff_col = f"a{eq_idx}{var_idx}"
                if coeff_col in row.index:
                    value = row[coeff_col]
                    if pd.isna(value):
                        coefficients.append("0")
                    else:
                        coefficients.append(str(value).strip())
                else:
                    coefficients.append("0")
            
            # Add constant term
            const_col = f"c{eq_idx}"
            if const_col in row.index:
                value = row[const_col]
                if pd.isna(value):
                    coefficients.append("0")
                else:
                    coefficients.append(str(value).strip())
            else:
                coefficients.append("0")
            
            # Join coefficients with comma
            equation_str = ",".join(coefficients)
            equations.append(equation_str)
        
        return equations
    
    def process_excel_equations(self, file_path: str, operation: str, version: str = "fx799", 
                              progress_callback: callable = None) -> Tuple[int, int, str]:
        """Process Excel file with equation data"""
        try:
            from equation_core import EquationProcessor
            
            # Check if large file processing needed
            is_large, file_info = self.is_large_file(file_path)
            
            if is_large:
                return self._process_large_excel_equations(file_path, operation, version, progress_callback)
            
            # Normal processing
            df = self.read_excel_data(file_path)
            
            # Validate structure
            is_valid, missing_cols = self.validate_excel_structure(df, operation)
            if not is_valid:
                raise Exception(f"Cấu trúc Excel không hợp lệ. Thiếu cột: {', '.join(missing_cols)}")
            
            # Initialize equation processor
            equation_processor = EquationProcessor()
            
            # Process each row
            results = []
            processed_count = 0
            error_count = 0
            
            for index, row in df.iterrows():
                try:
                    # Extract equation data from row
                    equations = self.extract_equation_data(row, operation)
                    
                    # Check if row has data
                    if not any(eq.replace(',', '').replace('0', '').strip() for eq in equations):
                        results.append("")  # Empty result for empty rows
                        continue
                    
                    # Process equation
                    result = equation_processor.process_single(operation, equations, version)
                    
                    if result.get('success'):
                        results.append(result['keylog'])
                        processed_count += 1
                    else:
                        results.append(f"ERROR: {result.get('error', 'Unknown error')}")
                        error_count += 1
                    
                    # Call progress callback if provided
                    if progress_callback:
                        progress_callback(index + 1, len(df))
                
                except Exception as e:
                    results.append(f"ERROR: {str(e)}")
                    error_count += 1
            
            # Create output file with results
            output_path = self.export_equation_results(df, results, operation, version)
            
            return processed_count, error_count, output_path
            
        except Exception as e:
            raise Exception(f"Lỗi xử lý Excel equations: {str(e)}")
    
    def _process_large_excel_equations(self, file_path: str, operation: str, version: str,
                                     progress_callback: callable = None) -> Tuple[int, int, str]:
        """Process large Excel files using chunked processing"""
        try:
            from equation_core import EquationProcessor
            
            print(f"🔥 Processing large file: {os.path.basename(file_path)}")
            
            # Initialize
            equation_processor = EquationProcessor()
            all_results = []
            processed_count = 0
            error_count = 0
            total_rows = 0
            
            # Determine optimal chunk size
            _, file_info = self.is_large_file(file_path)
            chunk_size = file_info.get('recommended_chunk_size', 500)
            
            # Process in chunks
            original_df = None
            for chunk_idx, chunk_df in enumerate(self.read_excel_data_chunked(file_path, chunk_size)):
                # Store first chunk structure for final export
                if original_df is None:
                    original_df = chunk_df.copy()
                
                # Validate chunk structure
                is_valid, missing_cols = self.validate_excel_structure(chunk_df, operation)
                if not is_valid:
                    raise Exception(f"Chunk {chunk_idx}: Cấu trúc Excel không hợp lệ. Thiếu cột: {', '.join(missing_cols)}")
                
                # Process chunk
                chunk_results = []
                for index, row in chunk_df.iterrows():
                    try:
                        equations = self.extract_equation_data(row, operation)
                        
                        if not any(eq.replace(',', '').replace('0', '').strip() for eq in equations):
                            chunk_results.append("")
                            continue
                        
                        result = equation_processor.process_single(operation, equations, version)
                        
                        if result.get('success'):
                            chunk_results.append(result['keylog'])
                            processed_count += 1
                        else:
                            chunk_results.append(f"ERROR: {result.get('error', 'Unknown error')}")
                            error_count += 1
                            
                    except Exception as e:
                        chunk_results.append(f"ERROR: {str(e)}")
                        error_count += 1
                
                all_results.extend(chunk_results)
                total_rows += len(chunk_df)
                
                # Progress callback
                if progress_callback:
                    progress_callback(total_rows, file_info.get('estimated_rows', total_rows))
                
                print(f"Processed chunk {chunk_idx + 1}, rows: {len(chunk_df)}")
            
            # Create combined results DataFrame for export
            if original_df is not None:
                # Read original file normally for export structure (risky for very large files)
                try:
                    full_df = pd.read_excel(file_path)
                    full_df.columns = full_df.columns.str.strip()
                except:
                    # Fallback: use reconstructed structure
                    print("Warning: Using reconstructed DataFrame for export")
                    full_df = original_df
                
                output_path = self.export_equation_results(full_df, all_results, operation, version)
            else:
                raise Exception("Không thể xử lý file - không có dữ liệu hợp lệ")
            
            return processed_count, error_count, output_path
            
        except Exception as e:
            raise Exception(f"Lỗi xử lý large file: {str(e)}")
    
    def export_equation_results(self, original_df: pd.DataFrame, encoded_results: List[str], 
                               operation: str, version: str) -> str:
        """Export equation results with Excel formatting"""
        try:
            result_df = original_df.copy()
            
            # Ensure results list matches DataFrame length
            if len(encoded_results) != len(result_df):
                # Pad or truncate results to match DataFrame
                if len(encoded_results) < len(result_df):
                    encoded_results.extend([''] * (len(result_df) - len(encoded_results)))
                else:
                    encoded_results = encoded_results[:len(result_df)]
            
            # Add results columns
            result_df['Keylog_Result'] = encoded_results
            result_df['Operation'] = operation
            result_df['Version'] = version
            result_df['Processed_Time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Generate output filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"equation_results_{timestamp}.xlsx"
            output_path = os.path.join(tempfile.gettempdir(), output_filename)
            
            # Export with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Results')
                
                # Format the worksheet
                worksheet = writer.sheets['Results']
                self._format_results_worksheet(worksheet, result_df)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Không thể xuất file kết quả: {str(e)}")
    
    def _format_results_worksheet(self, worksheet, df):
        """Format Excel worksheet with colors and fonts"""
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
            
            # Data formatting
            data_font = Font(name='Arial', size=10)
            result_font = Font(name='Arial', size=10, bold=True, color='2E7D32')
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # Find keylog column index
            keylog_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if 'Keylog_Result' in str(col_name):
                    keylog_col_idx = idx
                    break
            
            # Apply data formatting - limit to first 1000 rows for performance
            max_format_rows = min(len(df) + 2, 1000)
            for row in range(2, max_format_rows):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Special formatting for result column
                    if keylog_col_idx is not None and col == keylog_col_idx + 1:
                        cell.font = result_font
                    else:
                        cell.font = data_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Check only first 50 rows for performance
                for i, cell in enumerate(column):
                    if i > 50:
                        break
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Warning: Could not format worksheet: {e}")
    
    def create_equation_template(self, operation: str, output_path: str) -> str:
        """Create Excel template for equation data input"""
        try:
            if operation not in self.mapping:
                raise Exception(f"Operation '{operation}' không được hỗ trợ")
            
            operation_config = self.mapping[operation]
            required_cols = operation_config['required_columns']
            
            # Create sample data
            template_data = {}
            for col in required_cols:
                template_data[col] = self._get_equation_sample_data(col)
            
            # Add keylog result column
            template_data['Keylog_Result'] = [''] * len(next(iter(template_data.values())))
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Export template
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Template')
                
                # Format template
                worksheet = writer.sheets['Template']
                self._format_template_worksheet(worksheet, df, operation)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Không thể tạo template: {str(e)}")
    
    def _get_equation_sample_data(self, column: str) -> List[str]:
        """Generate sample data for equation template"""
        if column.startswith('a'):
            # Coefficient columns - provide varied examples
            if column.endswith('1'):
                return ['1', '2', '0', '1']
            elif column.endswith('2'):
                return ['2', '-1', '1', '0']
            elif column.endswith('3'):
                return ['1', '1', '2', '-1']
            elif column.endswith('4'):
                return ['0', '1', '1', '1']
            else:
                return ['1', '2', '0', '1']
        elif column.startswith('c'):
            # Constant columns
            if column.endswith('1'):
                return ['5', '7', '3', '10']
            elif column.endswith('2'):
                return ['3', '1', '2', '5']
            elif column.endswith('3'):
                return ['4', '2', '1', '7']
            elif column.endswith('4'):
                return ['6', '3', '4', '8']
            else:
                return ['5', '7', '3', '10']
        return ['', '', '', '']
    
    def _format_template_worksheet(self, worksheet, df, operation):
        """Format template worksheet with instructions"""
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Warning: Could not format template: {e}")
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about Excel file"""
        try:
            # Check if large file first
            is_large, large_file_info = self.is_large_file(file_path)
            
            if is_large:
                # Use openpyxl for large file info
                return self._get_large_file_info(file_path, large_file_info)
            else:
                # Use pandas for normal files
                df = self.read_excel_data(file_path)
                file_name = os.path.basename(file_path)
                
                return {
                    'file_name': file_name,
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns),
                    'file_size': os.path.getsize(file_path),
                    'file_size_mb': os.path.getsize(file_path) / (1024 * 1024),
                    'is_large_file': False,
                    'first_few_rows': df.head(3).to_dict('records') if len(df) > 0 else []
                }
        except Exception as e:
            raise Exception(f"Không thể đọc thông tin file: {str(e)}")
    
    def _get_large_file_info(self, file_path: str, large_file_info: Dict) -> Dict[str, Any]:
        """Get file info for large files without loading full data"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = [str(cell) for cell in header_row if cell is not None]
            
            # Get sample rows (first 3 data rows)
            sample_rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
                if i >= 3:
                    break
                row_dict = {col: str(cell) if cell is not None else "" for col, cell in zip(columns, row)}
                sample_rows.append(row_dict)
            
            wb.close()
            
            return {
                'file_name': os.path.basename(file_path),
                'total_rows': large_file_info.get('estimated_rows', 0),
                'total_columns': len(columns),
                'columns': columns,
                'file_size': os.path.getsize(file_path),
                'file_size_mb': large_file_info.get('file_size_mb', 0),
                'is_large_file': True,
                'recommended_chunk_size': large_file_info.get('recommended_chunk_size', 500),
                'first_few_rows': sample_rows,
                'warning': 'File lớn - khuyến nghị dùng xử lý chunked'
            }
            
        except Exception as e:
            raise Exception(f"Không thể phân tích file lớn: {str(e)}")
    
    def validate_data_quality(self, df: pd.DataFrame, operation: str) -> Dict[str, Any]:
        """Validate data quality in Excel file for equations"""
        quality_info = {
            'valid': True,
            'total_rows': len(df),
            'rows_with_data': 0,
            'rows_with_errors': 0,
            'missing_columns': [],
            'data_issues': []
        }
        
        # Check structure first
        is_valid, missing_cols = self.validate_excel_structure(df, operation)
        if not is_valid:
            quality_info['valid'] = False
            quality_info['missing_columns'] = missing_cols
            return quality_info
        
        # Sample-based validation for performance (max 1000 rows)
        sample_size = min(1000, len(df))
        sample_df = df.head(sample_size) if len(df) > 1000 else df
        quality_info['is_sample_validation'] = len(df) > 1000
        quality_info['sample_size'] = sample_size
        
        # Check rows for data quality
        for row_index in range(len(sample_df)):
            row = sample_df.iloc[row_index]
            has_data = False
            row_issues = []
            
            # Extract equation data
            try:
                equations = self.extract_equation_data(row, operation)
                if any(eq.replace(',', '').replace('0', '').strip() for eq in equations):
                    has_data = True
                    
                    # Validate equation coefficients
                    for eq_idx, eq in enumerate(equations):
                        coeffs = eq.split(',')
                        for coeff_idx, coeff in enumerate(coeffs):
                            if coeff and coeff.strip() and not self._is_valid_number(coeff.strip()):
                                row_issues.append(f"Phương trình {eq_idx+1}, hệ số {coeff_idx+1}: '{coeff}' không phải số hợp lệ")
            except Exception as e:
                row_issues.append(f"Lỗi đọc dữ liệu: {str(e)}")
            
            if has_data:
                quality_info['rows_with_data'] += 1
            
            if row_issues:
                quality_info['rows_with_errors'] += 1
                if len(quality_info['data_issues']) < 10:  # Limit to first 10 errors
                    quality_info['data_issues'].append({
                        'row': row_index + 2,  # Excel row number (1-indexed + header)
                        'issues': row_issues[:3]  # Limit to first 3 issues per row
                    })
        
        return quality_info
    
    def _is_valid_number(self, value: str) -> bool:
        """Check if value is a valid number"""
        try:
            float(value)
            return True
        except:
            return False